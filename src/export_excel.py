# -*- coding: utf-8 -*-
"""Excel / 报告导出 (P2-③, 对标 Wind/Choice 一键导出研报)。

把组合优化、风险归因、候选公司画像、因子挖掘结果汇总成 xlsx 工作簿，
供用户下载后在 Excel / WPS 二次加工。依赖 openpyxl（项目已安装）。
"""
from __future__ import annotations

import logging
import os
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

_EXPORT_DIR = os.path.join("data", "exports")


def _ensure_dir() -> str:
    os.makedirs(_EXPORT_DIR, exist_ok=True)
    return _EXPORT_DIR


def build_portfolio_report(
    symbols: List[str],
    objective: str = "max_sharpe",
    risk_parity_mode: bool = False,
    online: bool = False,
    window: int = 120,
    rf: float = 0.02,
) -> Dict[str, Any]:
    """汇总组合优化 + 风险归因 + 候选公司画像 + 当前最优因子。

    Returns: 结构化报告 dict（同时供 xlsx 与 JSON 复用）。
    """
    from src.portfolio_optimizer import optimize_portfolio, risk_attribution
    from src.storage import DatabaseManager, CompanyProfile, FactorMiningResult

    opt = optimize_portfolio(
        symbols=symbols, objective=objective, online=online,
        rf=rf, window=window, risk_parity_mode=risk_parity_mode,
    )
    if "error" in opt:
        return {"error": opt["error"], "symbols": symbols}

    # 候选公司画像
    # 注意: ORM 属性必须在 session 作用域内取出，否则 session 关闭后访问
    # r.code 会触发惰性刷新并抛出 DetachedInstanceError。因此 company 字典
    # 的构建放在 with 块内部完成。
    companies: List[Dict[str, Any]] = []
    factor_info: Optional[Dict[str, Any]] = None
    m = DatabaseManager.get_instance()
    with m.session_scope() as s:
        rows = s.query(CompanyProfile).filter(
            CompanyProfile.code.in_(symbols)).all() if hasattr(CompanyProfile, "code") else []
        by_code = {str(r.code): r for r in rows}
        # 当前最优因子
        frow = s.query(FactorMiningResult).filter_by(is_active=1).order_by(
            FactorMiningResult.ic.desc()).first()
        if frow:
            factor_info = {
                "factor_name": frow.factor_name,
                "factor_expr": frow.factor_expr,
                "ic": frow.ic,
                "sharpe": frow.sharpe,
            }

        for code in symbols:
            r = by_code.get(str(code))
            if r:
                companies.append({
                    "code": str(r.code),
                    "name": getattr(r, "name", "") or "",
                    "industry": getattr(r, "industry", "") or "",
                    "consensus_rating": getattr(r, "consensus_rating", None),
                    "target_price": getattr(r, "target_price", None),
                    "esg_rating": getattr(r, "esg_rating", None),
                    "esg_score": getattr(r, "esg_score", None),
                })
            else:
                companies.append({"code": code, "name": "", "industry": ""})

    return {
        "symbols": symbols,
        "method": opt.get("method"),
        "weights": opt.get("weights"),
        "exp_return": opt.get("exp_return"),
        "volatility": opt.get("volatility"),
        "sharpe": opt.get("sharpe"),
        "risk_attribution": opt.get("risk_attribution"),
        "companies": companies,
        "factor_info": factor_info,
        "objective": objective,
        "risk_parity_mode": risk_parity_mode,
        "online": online,
    }


def export_report_xlsx(report: Dict[str, Any], filename: Optional[str] = None) -> str:
    """把报告写入 xlsx，返回文件路径。

    Sheets: 组合概览 / 风险归因 / 候选标的 / 因子信号。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    _ensure_dir()
    if not filename:
        from datetime import datetime
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"portfolio_report_{ts}.xlsx"
    path = os.path.join(_EXPORT_DIR, filename)

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="305496")
    head_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=13, color="305496")

    def _style_header(ws, row: int, ncols: int):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center")

    # ---- Sheet 1: 组合概览 ----
    ws = wb.active
    ws.title = "组合概览"
    ws["A1"] = "组合优化报告"
    ws["A1"].font = title_font
    ws["A2"] = f"方法: {report.get('method')}    " \
               f"预期年化收益: {report.get('exp_return')}    " \
               f"年化波动: {report.get('volatility')}    " \
               f"夏普: {report.get('sharpe')}"
    ws["A2"].font = Font(bold=True)
    hdr = ["代码", "权重", "预期收益", "波动", "夏普"]
    ws.append([])
    ws.append(hdr)
    _style_header(ws, ws.max_row, len(hdr))
    symbols = report.get("symbols", [])
    weights = report.get("weights", [])
    for i, code in enumerate(symbols):
        w = weights[i] if i < len(weights) else None
        ws.append([code, w, None, None, None])
    # 把权重格式化为百分比
    for r in range(5, 5 + len(symbols)):
        ws.cell(row=r, column=2).number_format = "0.00%"

    # ---- Sheet 2: 风险归因 ----
    ws2 = wb.create_sheet("风险归因")
    ws2["A1"] = "风险归因（波动分解到每只标的）"
    ws2["A1"].font = title_font
    attr = report.get("risk_attribution", {})
    per = attr.get("per_asset", [])
    hdr2 = ["代码", "权重", "边际风险(MCR)", "成分风险(CCR)", "百分比贡献"]
    ws2.append([])
    ws2.append(hdr2)
    _style_header(ws2, ws2.max_row, len(hdr2))
    for i, code in enumerate(symbols):
        a = per[i] if i < len(per) else {}
        ws2.append([
            code, a.get("weight"), a.get("mcr"), a.get("ccr"), a.get("pct"),
        ])
    for r in range(4, 4 + len(symbols)):
        ws2.cell(row=r, column=2).number_format = "0.00%"
        ws2.cell(row=r, column=5).number_format = "0.00%"
    last = 4 + len(symbols)
    ws2.cell(row=last + 1, column=1, value="组合波动").font = Font(bold=True)
    ws2.cell(row=last + 1, column=2, value=attr.get("portfolio_vol"))
    ws2.cell(row=last + 2, column=1, value="风险集中度 HHI").font = Font(bold=True)
    ws2.cell(row=last + 2, column=2, value=attr.get("hhi_risk"))

    # ---- Sheet 3: 候选标的 ----
    ws3 = wb.create_sheet("候选标的")
    ws3["A1"] = "候选标的基本面画像"
    ws3["A1"].font = title_font
    hdr3 = ["代码", "名称", "行业", "一致预期评级", "目标价", "ESG评级", "ESG评分"]
    ws3.append([])
    ws3.append(hdr3)
    _style_header(ws3, ws3.max_row, len(hdr3))
    for c in report.get("companies", []):
        ws3.append([
            c.get("code"), c.get("name"), c.get("industry"),
            c.get("consensus_rating"), c.get("target_price"),
            c.get("esg_rating"), c.get("esg_score"),
        ])

    # ---- Sheet 4: 因子信号 ----
    ws4 = wb.create_sheet("因子信号")
    ws4["A1"] = "当前最优挖掘因子"
    ws4["A1"].font = title_font
    finfo = report.get("factor_info")
    if finfo:
        ws4.append([])
        ws4.append(["因子名", finfo.get("factor_name")])
        ws4.append(["表达式", finfo.get("factor_expr")])
        ws4.append(["IC", finfo.get("ic")])
        ws4.append(["夏普", finfo.get("sharpe")])
    else:
        ws4.append([])
        ws4.append(["（暂无激活的挖掘因子）"])

    wb.save(path)
    return path
